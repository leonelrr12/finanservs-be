tabla = "profesions_acp"
host = "localhost"
user = "root"
password = ""
database = "finanservs"

#********************************
# Dockert
#********************************
# host='localhost',
# database='finanservs',
# user='rsanchez',
# password='cafekotowa'
#********************************
 
# import pandas as pd
# xls = pd.read_excel(fichero)
# print(xls.columns)
# values = xls['Titulo'].values
# print(values)

# import pandas as pd
# xls = pd.ExcelFile(fichero)
# print(xls.sheet_names)
# df = xls.parse('Original')
# print(df)

# # MySql de James - IS
# cnn = mysql.connector.connect(
#     host="69.10.63.218",
#     database="finanservs",
#     user="AdminFinanservs",
#     password="0t_pYv70"     
# )



import mysql.connector

def insert_acp(id, titulo, grupo, segmento, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO profesions_acp (id,titulo,grupo,segmento) VALUES (%s,%s,%s,%s)"

        record = (id,titulo,grupo,segmento)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_lw(id, titulo, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO profesions_lw (id,titulo) VALUES (%s,%s)"

        record = (id,titulo)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_inst(id, name, subgrupo, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO institutions (id,name,subgrupo) VALUES (%s,%s,%s)"

        record = (id,name,subgrupo)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_plan(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO planillas_j (id,name) VALUES (%s,%s)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_pais(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO nationality (id,name,is_active) VALUES (%s,%s,1)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_prov(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO provinces (id,name) VALUES (%s,%s)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_dist(id, idProv, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO districts (id, idProv, name) VALUES (%s,%s,%s)"

        record = (id,idProv,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_corr(id, idDist, idProv, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO counties (id,  idProv, idDist, name) VALUES (%s,%s,%s,%s)"

        record = (id,idDist,idProv,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def planilla_css(cedula, nombre, entidad, cargo, salario, inicio, contrato, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO planilla_css (cedula, nombre_completo, entidad, cargo, salario_mensual, inicio_labores, contrato) VALUES (%s,%s,%s,%s,%s,%s,%s)"

        record = (cedula, nombre, entidad, cargo, salario, inicio, contrato)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def planilla_contraloria(cedula, nombre, entidad, cargo, salario, gasto, inicio, contrato, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO planilla_css (cedula, nombre_completo, entidad, cargo, salario_mensual, gasto_mensual, inicio_labores, contrato) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"

        record = (cedula, nombre, entidad, cargo, salario, gasto, inicio, contrato)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


# # librerias para correos
# import smtplib, ssl
# import getpass
# libresia para excel
import openpyxl

# # credenciales de la cuenta de correo
# username = input("Ingrese su nombre de usuario: ")
# password = getpass.getpass("Ingrese su password: ")

# # Crear la conexion
# context = ssl.create_default_context()

# with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
#     server.login(username, password)
#     print("Inició sesión")

#     destino = "email"
#     mensaje = "Cuerpo del correo"
#     server.sendmail(username, destino, mensaje)


def fcnn():
    # return mysql.connector.connect(
    # host="134.122.114.209",
    # database="finanservs",
    # user="dbkotowa",
    # password="lr1wapia8dem9cla",
    # port="3306"

    mydb = mysql.connector.connect(
        host="db-mysql-nyc3-04501-do-user-7220305-0.b.db.ondigitalocean.com",
        user="dbkotowa",
        password="FtktBOEQa9UHFDPe",
        database="finanservs",
        port="25060"
    )

    print(mydb)

    if(mydb):
        print("Connection Successful")
    else:
        print("Connection is fail")

    return mydb



## No. 1
def acp():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\Profesiones_ACP.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'D751']

    # MySql de Digital Ocean
    cnn = fcnn()

    print("Profsionales ACP ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_acp(data[0], data[1], data[2], data[3], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 2
def lw():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\Profesiones_Linea_Blanca.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A5' : 'B105']
    cnn = fcnn()

    print("Linea Blanca ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_lw(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 3
def inst():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\Instituciones.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A3' : 'C70']

    cnn = fcnn()

    print("Instituciones ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_inst(data[0], data[1], data[2], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 4
def plan():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\Planillas_Jubilados.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A3' : 'B15']

    cnn = fcnn()

    print("Planillas Jubilados ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_plan(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 567
def pais():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\paises_estandar.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'B238']

    cnn = fcnn()

    print("Paises ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_pais(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")


## No. 5
def prov():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\prov-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'B11']

    cnn = fcnn()

    print("Provincias ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_prov(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 6
def dist():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\dist-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'C71']

    cnn = fcnn()

    print("Distritos ...") 
    for fila in celdas:
        data = [celda.value for celda in fila]
        # print(data)
        insert_dist(data[0], data[1], data[2], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 7
def corr():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\corr-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'D599']

    cnn = fcnn()

    print("Corregimientos ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_corr(data[0], data[1], data[2], data[3], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")


## No. 8
def css():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\planilla-institucional-de-la-css-febrero-2022.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active
    max_row = hoja.max_row
    
    celdas = hoja['A4' : 'H'+str(max_row)]  

    cnn = fcnn()
    mycursor = cnn.cursor()

    query = "delete from planilla_css where entidad = 'CAJA DE SEGURO SOCIAL'"
    mycursor.execute(query)


    print("Planilla CSS ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        # print(data)
        planilla_css(data[2], data[1], 'CAJA DE SEGURO SOCIAL', data[3], data[4], data[5], data[7], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

## No. 9
def contraloria():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\otros\contraloria.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active
    max_row = hoja.max_row

    celdas = hoja['A6' : 'H'+str(max_row)]  

    cnn = fcnn()
    mycursor = cnn.cursor()

    query = "delete from planilla_css where entidad <> 'CAJA DE SEGURO SOCIAL'"
    mycursor.execute(query)


    print("Planillas de Contraloria ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        nombreCompleto = data[0]+' '+data[1]
        f1 = data[7].split('/')
        inicio=f1[2]+'-'+f1[1]+'-'+f1[0]
        f1 = data[2].split('-')
        cedula = f1[0]+'-'+str(int(f1[1]))+'-'+str(int(f1[2]))
        planilla_contraloria(cedula, nombreCompleto, 'TRIBUNAL DE CUENTAS', data[3], data[4], data[5], inicio, data[6], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")


#acp()
#lw()
#inst()
#plan()
#pais()
#prov()
#dist()
#corr()
# css()
contraloria()
